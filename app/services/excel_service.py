"""Exportación / importación en Excel de las tablas maestras.

Cuatro entidades, una hoja por fichero:
    - empresas    (clave única: CIF)
    - centros     (clave única: empresa_cif + codigo_centro / nombre)
    - empleados   (clave única: NIF)
    - usuarios    (clave única: DNI)

El import es transaccional y "upsert": si la clave única ya existe, se
actualiza; si no, se inserta. Cualquier fila inválida aborta el commit.
"""

from __future__ import annotations

import difflib
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.models import (
    CentroHorarioApertura, CentroTrabajo, DotacionCentro, Empleado, Empresa,
    PlantillaTurno, Usuario,
)
from app.models.usuario import ROL_ADMIN, ROL_EMPLEADO, ROLES_VALIDOS


# --- Constantes ---------------------------------------------------------

# Mapeo entre código de día castellano y entero 0..6 (lun..dom).
_DIA_CODIGOS = ["L", "M", "X", "J", "V", "S", "D"]
_DIA_A_INT = {c: i for i, c in enumerate(_DIA_CODIGOS)}

_BOOL_TRUE = {"si", "sí", "s", "yes", "y", "true", "1", "verdadero", "v"}
_BOOL_FALSE = {"no", "n", "false", "0", "falso", "f"}


# --- Resultado ----------------------------------------------------------

@dataclass
class ImportResult:
    creados: int = 0
    actualizados: int = 0
    # Fallos estructurales: abortan el commit (fichero ilegible, columnas, etc.).
    errores: list[tuple[int, str]] = field(default_factory=list)
    # Incidencias por fila que NO impiden el commit (import parcial): tiendas
    # no emparejadas, ambiguas, horarios inválidos, emparejamientos dudosos.
    avisos: list[tuple[int, str]] = field(default_factory=list)
    commit_ok: bool = False

    @property
    def total_filas_validas(self) -> int:
        return self.creados + self.actualizados


# --- Helpers de parseo --------------------------------------------------

def _norm_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (int, float)):
        # Evita "1234.0" cuando Excel guarda un código como número.
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    return str(v).strip()


def _norm_upper(v: Any) -> str:
    return _norm_str(v).upper()


def _norm_lower(v: Any) -> str:
    return _norm_str(v).lower()


def _to_bool(v: Any, default: bool | None = None) -> bool | None:
    """Devuelve True/False según el valor, o `default` si la celda está vacía."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip().lower()
    if s in _BOOL_TRUE:
        return True
    if s in _BOOL_FALSE:
        return False
    raise ValueError(f"Valor booleano no reconocido: {v!r}")


def _bool_to_excel(v: bool | None) -> str:
    if v is True:
        return "Sí"
    if v is False:
        return "No"
    return ""


def _to_date(v: Any) -> date | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    return datetime.strptime(s, "%Y-%m-%d").date()


def _to_horario(v: Any) -> str | None:
    """Cadena HH:MM o None. Acepta también un datetime.time de Excel."""
    if v is None:
        return None
    if hasattr(v, "hour") and hasattr(v, "minute"):
        return f"{v.hour:02d}:{v.minute:02d}"
    s = _norm_str(v)
    if not s:
        return None
    # Validación mínima: tiene que parecer HH:MM
    try:
        h, m = s.split(":")
        return f"{int(h):02d}:{int(m):02d}"
    except Exception as exc:
        raise ValueError(f"Horario no reconocido: {v!r} (esperado HH:MM)") from exc


def _to_numero(v: Any) -> float | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return float(str(v).strip().replace(",", "."))


def _dias_to_str(dias: list[int] | None) -> str:
    if not dias:
        return ""
    return ",".join(_DIA_CODIGOS[i] for i in sorted(set(dias)) if 0 <= i <= 6)


def _to_dow(v: Any) -> int:
    """Convierte un valor a día de semana 0..6 (acepta L,M,X,J,V,S,D o 0-6)."""
    s = _norm_upper(v)
    if not s:
        raise ValueError("dia_semana es obligatorio.")
    if s.isdigit():
        i = int(s)
        if 0 <= i <= 6:
            return i
        raise ValueError(f"dia_semana fuera de rango: {s}")
    if s in _DIA_A_INT:
        return _DIA_A_INT[s]
    raise ValueError(f"dia_semana no reconocido: {v!r} (use L,M,X,J,V,S,D o 0-6).")


def _str_to_dias(v: Any) -> list[int] | None:
    s = _norm_upper(v)
    if not s:
        return None
    out: set[int] = set()
    for token in s.replace(";", ",").split(","):
        t = token.strip()
        if not t:
            continue
        if t.isdigit():
            i = int(t)
            if 0 <= i <= 6:
                out.add(i)
                continue
            raise ValueError(f"Día fuera de rango: {t}")
        if t in _DIA_A_INT:
            out.add(_DIA_A_INT[t])
            continue
        raise ValueError(f"Día no reconocido: {t!r}")
    return sorted(out) or None


# --- Helpers de export --------------------------------------------------

def _crear_libro(hoja: str, cabeceras: list[str]) -> tuple[Workbook, Any]:
    wb = Workbook()
    ws = wb.active
    ws.title = hoja
    ws.append(cabeceras)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    return wb, ws


def _autosize(ws: Any, datos: Iterable[list[Any]], cabeceras: list[str]) -> None:
    anchos = [len(h) for h in cabeceras]
    for fila in datos:
        for i, valor in enumerate(fila):
            s = "" if valor is None else str(valor)
            if len(s) > anchos[i]:
                anchos[i] = len(s)
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(ancho + 2, 60)


def _serializar(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _leer_filas(file_bytes: bytes, cabeceras_esperadas: list[str]) -> tuple[list[dict], list[tuple[int, str]]]:
    """Devuelve (filas, errores). Cada fila es un dict cabecera→valor (None si vacía).

    `errores` solo se rellena en problemas estructurales (cabecera ausente).
    Filas completamente vacías se descartan en silencio.
    """
    errores: list[tuple[int, str]] = []
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        errores.append((0, f"No se pudo abrir el fichero Excel: {exc}"))
        return [], errores
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        cabeceras = [c for c in next(rows)]
    except StopIteration:
        errores.append((0, "El fichero está vacío."))
        return [], errores

    cabeceras_norm = [(c or "").strip().lower() if isinstance(c, str) else c for c in cabeceras]
    faltan = [c for c in cabeceras_esperadas if c not in cabeceras_norm]
    if faltan:
        errores.append((1, f"Faltan columnas obligatorias: {', '.join(faltan)}"))
        return [], errores

    filas: list[dict] = []
    for idx, row in enumerate(rows, start=2):
        if row is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        d = {cabeceras_norm[i]: row[i] if i < len(row) else None for i in range(len(cabeceras_norm))}
        d["__row__"] = idx
        filas.append(d)
    return filas, errores


# --- Empresas -----------------------------------------------------------

_EMPRESA_CABECERAS = ["id", "nombre", "cif", "direccion", "codigo_cuenta_cotizacion", "activa"]


def exportar_empresas(db: Session) -> bytes:
    empresas = db.query(Empresa).order_by(Empresa.nombre).all()
    wb, ws = _crear_libro("Empresas", _EMPRESA_CABECERAS)
    datos = []
    for e in empresas:
        fila = [
            e.id,
            e.nombre,
            e.cif,
            e.direccion or "",
            e.codigo_cuenta_cotizacion or "",
            _bool_to_excel(e.activa),
        ]
        ws.append(fila)
        datos.append(fila)
    _autosize(ws, datos, _EMPRESA_CABECERAS)
    return _serializar(wb)


def importar_empresas(db: Session, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    filas, errores = _leer_filas(file_bytes, ["nombre", "cif"])
    result.errores.extend(errores)
    if errores:
        return result

    cifs_en_fichero: set[str] = set()
    pendientes: list[tuple[Empresa | None, dict, int]] = []

    for fila in filas:
        n = fila["__row__"]
        try:
            cif = _norm_upper(fila.get("cif"))
            if not cif:
                raise ValueError("El CIF es obligatorio.")
            if cif in cifs_en_fichero:
                raise ValueError(f"CIF duplicado en el fichero: {cif}")
            cifs_en_fichero.add(cif)
            nombre = _norm_str(fila.get("nombre"))
            if not nombre:
                raise ValueError("El nombre es obligatorio.")
            existente = db.query(Empresa).filter(Empresa.cif == cif).one_or_none()
            pendientes.append((existente, fila, n))
        except ValueError as exc:
            result.errores.append((n, str(exc)))

    if result.errores:
        return result

    for existente, fila, n in pendientes:
        cif = _norm_upper(fila.get("cif"))
        nombre = _norm_str(fila.get("nombre"))
        direccion = _norm_str(fila.get("direccion")) or None
        ccc = _norm_str(fila.get("codigo_cuenta_cotizacion")) or None
        try:
            activa = _to_bool(fila.get("activa"), default=True if existente is None else existente.activa)
        except ValueError as exc:
            result.errores.append((n, f"Columna 'activa': {exc}"))
            continue
        if existente is None:
            db.add(Empresa(
                nombre=nombre, cif=cif, direccion=direccion,
                codigo_cuenta_cotizacion=ccc, activa=bool(activa),
            ))
            result.creados += 1
        else:
            existente.nombre = nombre
            existente.direccion = direccion
            existente.codigo_cuenta_cotizacion = ccc
            existente.activa = bool(activa)
            result.actualizados += 1

    if result.errores:
        db.rollback()
        return result
    db.commit()
    result.commit_ok = True
    return result


# --- Centros ------------------------------------------------------------

_CENTRO_CABECERAS = [
    "id", "empresa_cif", "nombre", "direccion", "ciudad", "provincia",
    "codigo_postal", "codigo_centro", "activo",
]


def exportar_centros(db: Session) -> bytes:
    centros = (
        db.query(CentroTrabajo).join(Empresa)
        .order_by(Empresa.nombre, CentroTrabajo.nombre).all()
    )
    wb, ws = _crear_libro("Centros", _CENTRO_CABECERAS)
    datos = []
    for c in centros:
        fila = [
            c.id,
            c.empresa.cif if c.empresa else "",
            c.nombre,
            c.direccion or "",
            c.ciudad or "",
            c.provincia or "",
            c.codigo_postal or "",
            c.codigo_centro or "",
            _bool_to_excel(c.activo),
        ]
        ws.append(fila)
        datos.append(fila)
    _autosize(ws, datos, _CENTRO_CABECERAS)
    return _serializar(wb)


def _buscar_centro(db: Session, empresa_id: int, codigo: str, nombre: str) -> CentroTrabajo | None:
    if codigo:
        c = (
            db.query(CentroTrabajo)
            .filter(CentroTrabajo.empresa_id == empresa_id, CentroTrabajo.codigo_centro == codigo)
            .one_or_none()
        )
        if c is not None:
            return c
    return (
        db.query(CentroTrabajo)
        .filter(CentroTrabajo.empresa_id == empresa_id, CentroTrabajo.nombre == nombre)
        .first()
    )


def importar_centros(db: Session, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    filas, errores = _leer_filas(file_bytes, ["empresa_cif", "nombre"])
    result.errores.extend(errores)
    if errores:
        return result

    claves_en_fichero: set[tuple[str, str]] = set()
    for fila in filas:
        n = fila["__row__"]
        try:
            empresa_cif = _norm_upper(fila.get("empresa_cif"))
            if not empresa_cif:
                raise ValueError("empresa_cif es obligatorio.")
            empresa = db.query(Empresa).filter(Empresa.cif == empresa_cif).one_or_none()
            if empresa is None:
                raise ValueError(f"No existe la empresa con CIF {empresa_cif}.")
            nombre = _norm_str(fila.get("nombre"))
            if not nombre:
                raise ValueError("El nombre es obligatorio.")
            codigo = _norm_str(fila.get("codigo_centro"))
            clave = (empresa_cif, codigo or nombre.lower())
            if clave in claves_en_fichero:
                raise ValueError(f"Centro duplicado en el fichero: {empresa_cif} / {codigo or nombre}")
            claves_en_fichero.add(clave)
            existente = _buscar_centro(db, empresa.id, codigo, nombre)
            try:
                activo = _to_bool(
                    fila.get("activo"),
                    default=True if existente is None else existente.activo,
                )
            except ValueError as exc:
                raise ValueError(f"Columna 'activo': {exc}")
            if existente is None:
                db.add(CentroTrabajo(
                    empresa_id=empresa.id,
                    nombre=nombre,
                    direccion=_norm_str(fila.get("direccion")) or None,
                    ciudad=_norm_str(fila.get("ciudad")) or None,
                    provincia=_norm_str(fila.get("provincia")) or None,
                    codigo_postal=_norm_str(fila.get("codigo_postal")) or None,
                    codigo_centro=codigo or None,
                    activo=bool(activo),
                ))
                result.creados += 1
            else:
                existente.empresa_id = empresa.id
                existente.nombre = nombre
                existente.direccion = _norm_str(fila.get("direccion")) or None
                existente.ciudad = _norm_str(fila.get("ciudad")) or None
                existente.provincia = _norm_str(fila.get("provincia")) or None
                existente.codigo_postal = _norm_str(fila.get("codigo_postal")) or None
                existente.codigo_centro = codigo or None
                existente.activo = bool(activo)
                result.actualizados += 1
        except ValueError as exc:
            result.errores.append((n, str(exc)))

    if result.errores:
        db.rollback()
        return result
    db.commit()
    result.commit_ok = True
    return result


# --- Empleados ----------------------------------------------------------

_EMPLEADO_CABECERAS = [
    "id", "empresa_cif", "centro_codigo_o_nombre", "nombre", "apellidos", "nif",
    "numero_afiliacion_ss", "categoria_profesional", "tipo_contrato",
    "horas_semanales", "fecha_alta", "fecha_baja", "activo", "email", "telefono",
    "inicio_jornada", "inicio_pausa", "fin_pausa", "fin_jornada",
    "dias_laborables_habituales",
    "admite_horas_extras", "disponible_desplazamiento", "dias_vacaciones_anuales",
    "centros_cubribles",
]


def _centro_referencia(c: CentroTrabajo | None) -> str:
    if c is None:
        return ""
    return c.codigo_centro or c.nombre


def exportar_empleados(db: Session) -> bytes:
    empleados = (
        db.query(Empleado).order_by(Empleado.apellidos, Empleado.nombre).all()
    )
    wb, ws = _crear_libro("Empleados", _EMPLEADO_CABECERAS)
    datos = []
    for e in empleados:
        fila = [
            e.id,
            e.empresa.cif if e.empresa else "",
            _centro_referencia(e.centro_trabajo),
            e.nombre,
            e.apellidos,
            e.nif,
            e.numero_afiliacion_ss or "",
            e.categoria_profesional or "",
            e.tipo_contrato or "",
            float(e.horas_semanales) if e.horas_semanales is not None else None,
            e.fecha_alta.isoformat() if e.fecha_alta else "",
            e.fecha_baja.isoformat() if e.fecha_baja else "",
            _bool_to_excel(e.activo),
            e.email or "",
            e.telefono or "",
            e.inicio_jornada or "",
            e.inicio_pausa or "",
            e.fin_pausa or "",
            e.fin_jornada or "",
            _dias_to_str(e.dias_laborables_habituales),
            _bool_to_excel(e.admite_horas_extras),
            _bool_to_excel(e.disponible_desplazamiento),
            e.dias_vacaciones_anuales,
            ", ".join(_centro_referencia(c) for c in e.centros_cubribles),
        ]
        ws.append(fila)
        datos.append(fila)
    _autosize(ws, datos, _EMPLEADO_CABECERAS)
    return _serializar(wb)


def _resolver_centro(db: Session, empresa_id: int, referencia: str) -> CentroTrabajo:
    """Resuelve un centro por código (preferente) o nombre dentro de una empresa.

    Lanza ValueError si no se encuentra o si el nombre es ambiguo.
    """
    if not referencia:
        raise ValueError("centro_codigo_o_nombre es obligatorio.")
    por_codigo = (
        db.query(CentroTrabajo)
        .filter(CentroTrabajo.empresa_id == empresa_id, CentroTrabajo.codigo_centro == referencia)
        .all()
    )
    if len(por_codigo) == 1:
        return por_codigo[0]
    if len(por_codigo) > 1:
        raise ValueError(f"Hay varios centros con código {referencia} en la empresa.")
    por_nombre = (
        db.query(CentroTrabajo)
        .filter(CentroTrabajo.empresa_id == empresa_id, CentroTrabajo.nombre == referencia)
        .all()
    )
    if len(por_nombre) == 1:
        return por_nombre[0]
    if len(por_nombre) > 1:
        raise ValueError(f"Hay varios centros con nombre {referencia!r} en la empresa.")
    raise ValueError(f"No se encontró el centro {referencia!r} en la empresa.")


def _resolver_cubribles(
    db: Session, empresa_id: int, valor: Any, centro_propio_id: int,
) -> list[CentroTrabajo]:
    """Resuelve la lista (separada por comas) de centros cubribles.

    Excluye el centro propio del empleado. Lanza ValueError si alguna
    referencia no se encuentra.
    """
    s = _norm_str(valor)
    if not s:
        return []
    out: list[CentroTrabajo] = []
    vistos: set[int] = set()
    for token in s.replace(";", ",").split(","):
        ref = token.strip()
        if not ref:
            continue
        centro = _resolver_centro(db, empresa_id, ref)
        if centro.id != centro_propio_id and centro.id not in vistos:
            out.append(centro)
            vistos.add(centro.id)
    return out


def importar_empleados(db: Session, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    filas, errores = _leer_filas(
        file_bytes,
        ["empresa_cif", "centro_codigo_o_nombre", "nombre", "apellidos", "nif"],
    )
    result.errores.extend(errores)
    if errores:
        return result

    nifs_en_fichero: set[str] = set()
    for fila in filas:
        n = fila["__row__"]
        try:
            nif = _norm_upper(fila.get("nif"))
            if not nif:
                raise ValueError("El NIF es obligatorio.")
            if nif in nifs_en_fichero:
                raise ValueError(f"NIF duplicado en el fichero: {nif}")
            nifs_en_fichero.add(nif)
            empresa_cif = _norm_upper(fila.get("empresa_cif"))
            empresa = db.query(Empresa).filter(Empresa.cif == empresa_cif).one_or_none()
            if empresa is None:
                raise ValueError(f"No existe la empresa con CIF {empresa_cif}.")
            referencia = _norm_str(fila.get("centro_codigo_o_nombre"))
            centro = _resolver_centro(db, empresa.id, referencia)
            nombre = _norm_str(fila.get("nombre"))
            apellidos = _norm_str(fila.get("apellidos"))
            if not nombre or not apellidos:
                raise ValueError("nombre y apellidos son obligatorios.")
            horas = _to_numero(fila.get("horas_semanales"))
            fecha_alta = _to_date(fila.get("fecha_alta"))
            fecha_baja = _to_date(fila.get("fecha_baja"))
            inicio_jornada = _to_horario(fila.get("inicio_jornada"))
            inicio_pausa = _to_horario(fila.get("inicio_pausa"))
            fin_pausa = _to_horario(fila.get("fin_pausa"))
            fin_jornada = _to_horario(fila.get("fin_jornada"))
            dias = _str_to_dias(fila.get("dias_laborables_habituales"))
            existente = db.query(Empleado).filter(Empleado.nif == nif).one_or_none()
            activo = _to_bool(
                fila.get("activo"),
                default=True if existente is None else existente.activo,
            )
            admite_he = _to_bool(
                fila.get("admite_horas_extras"),
                default=False if existente is None else existente.admite_horas_extras,
            )
            disp_desp = _to_bool(
                fila.get("disponible_desplazamiento"),
                default=False if existente is None else existente.disponible_desplazamiento,
            )
            dvac = _to_numero(fila.get("dias_vacaciones_anuales"))
            dvac = int(dvac) if dvac is not None else (
                existente.dias_vacaciones_anuales if existente is not None else 30
            )
            cubribles = _resolver_cubribles(
                db, empresa.id, fila.get("centros_cubribles"), centro.id,
            )
            if existente is None:
                nuevo = Empleado(
                    empresa_id=empresa.id,
                    centro_trabajo_id=centro.id,
                    nombre=nombre,
                    apellidos=apellidos,
                    nif=nif,
                    numero_afiliacion_ss=_norm_str(fila.get("numero_afiliacion_ss")) or None,
                    categoria_profesional=_norm_str(fila.get("categoria_profesional")) or None,
                    tipo_contrato=_norm_str(fila.get("tipo_contrato")) or None,
                    horas_semanales=horas if horas is not None else 40.0,
                    fecha_alta=fecha_alta,
                    fecha_baja=fecha_baja,
                    activo=bool(activo),
                    email=(_norm_lower(fila.get("email")) or None),
                    telefono=_norm_str(fila.get("telefono")) or None,
                    inicio_jornada=inicio_jornada,
                    inicio_pausa=inicio_pausa,
                    fin_pausa=fin_pausa,
                    fin_jornada=fin_jornada,
                    dias_laborables_habituales=dias,
                    admite_horas_extras=bool(admite_he),
                    disponible_desplazamiento=bool(disp_desp),
                    dias_vacaciones_anuales=dvac,
                )
                nuevo.centros_cubribles = cubribles
                db.add(nuevo)
                result.creados += 1
            else:
                existente.empresa_id = empresa.id
                existente.centro_trabajo_id = centro.id
                existente.nombre = nombre
                existente.apellidos = apellidos
                existente.nif = nif
                existente.numero_afiliacion_ss = _norm_str(fila.get("numero_afiliacion_ss")) or None
                existente.categoria_profesional = _norm_str(fila.get("categoria_profesional")) or None
                existente.tipo_contrato = _norm_str(fila.get("tipo_contrato")) or None
                if horas is not None:
                    existente.horas_semanales = horas
                existente.fecha_alta = fecha_alta
                existente.fecha_baja = fecha_baja
                existente.activo = bool(activo)
                existente.email = (_norm_lower(fila.get("email")) or None)
                existente.telefono = _norm_str(fila.get("telefono")) or None
                existente.inicio_jornada = inicio_jornada
                existente.inicio_pausa = inicio_pausa
                existente.fin_pausa = fin_pausa
                existente.fin_jornada = fin_jornada
                existente.dias_laborables_habituales = dias
                existente.admite_horas_extras = bool(admite_he)
                existente.disponible_desplazamiento = bool(disp_desp)
                existente.dias_vacaciones_anuales = dvac
                existente.centros_cubribles = cubribles
                result.actualizados += 1
        except ValueError as exc:
            result.errores.append((n, str(exc)))

    if result.errores:
        db.rollback()
        return result
    db.commit()
    result.commit_ok = True
    return result


# --- Usuarios -----------------------------------------------------------

_USUARIO_CABECERAS = ["id", "dni", "empleado_nif", "rol", "activo"]


def exportar_usuarios(db: Session) -> bytes:
    usuarios = (
        db.query(Usuario)
        .order_by(Usuario.es_admin.desc(), Usuario.dni).all()
    )
    wb, ws = _crear_libro("Usuarios", _USUARIO_CABECERAS)
    datos = []
    for u in usuarios:
        fila = [
            u.id,
            u.dni,
            u.empleado.nif if u.empleado else "",
            u.rol,
            _bool_to_excel(u.activo),
        ]
        ws.append(fila)
        datos.append(fila)
    _autosize(ws, datos, _USUARIO_CABECERAS)
    return _serializar(wb)


def importar_usuarios(db: Session, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    filas, errores = _leer_filas(file_bytes, ["dni"])
    result.errores.extend(errores)
    if errores:
        return result

    dnis_en_fichero: set[str] = set()
    empleado_nifs_usados: set[str] = set()

    for fila in filas:
        n = fila["__row__"]
        try:
            dni = _norm_upper(fila.get("dni"))
            if not dni:
                raise ValueError("El DNI es obligatorio.")
            if dni in dnis_en_fichero:
                raise ValueError(f"DNI duplicado en el fichero: {dni}")
            dnis_en_fichero.add(dni)
            empleado_nif = _norm_upper(fila.get("empleado_nif"))
            empleado: Empleado | None = None
            if empleado_nif:
                empleado = db.query(Empleado).filter(Empleado.nif == empleado_nif).one_or_none()
                if empleado is None:
                    raise ValueError(f"No existe el empleado con NIF {empleado_nif}.")
                if empleado_nif in empleado_nifs_usados:
                    raise ValueError(
                        f"El NIF {empleado_nif} aparece en más de una fila de usuarios."
                    )
                empleado_nifs_usados.add(empleado_nif)
            existente = db.query(Usuario).filter(Usuario.dni == dni).one_or_none()
            rol_raw = _norm_lower(fila.get("rol"))
            if rol_raw:
                if rol_raw not in ROLES_VALIDOS:
                    raise ValueError(
                        f"Rol no reconocido: {rol_raw!r}. Use admin, rrhh o empleado."
                    )
                rol = rol_raw
            else:
                # Compatibilidad con ficheros antiguos con columna `es_admin`.
                es_admin_legacy = _to_bool(fila.get("es_admin"), default=None)
                if es_admin_legacy is not None:
                    rol = ROL_ADMIN if es_admin_legacy else ROL_EMPLEADO
                elif existente is not None:
                    rol = existente.rol
                else:
                    rol = ROL_EMPLEADO
            activo = _to_bool(
                fila.get("activo"),
                default=True if existente is None else existente.activo,
            )
            # Si el empleado ya tiene OTRO usuario distinto al actual, no se puede asignar.
            if empleado is not None and empleado.usuario is not None:
                if existente is None or empleado.usuario.id != existente.id:
                    raise ValueError(
                        f"El empleado {empleado_nif} ya tiene otro usuario asignado."
                    )
            if existente is None:
                db.add(Usuario(
                    dni=dni,
                    empleado_id=empleado.id if empleado else None,
                    rol=rol,
                    activo=bool(activo),
                ))
                result.creados += 1
            else:
                existente.dni = dni
                existente.empleado_id = empleado.id if empleado else None
                existente.rol = rol
                existente.activo = bool(activo)
                result.actualizados += 1
        except ValueError as exc:
            result.errores.append((n, str(exc)))

    if result.errores:
        db.rollback()
        return result

    # Validación final: tras aplicar todos los cambios, debe seguir habiendo
    # al menos un admin activo en la BD (sin commit todavía).
    db.flush()
    admins_activos = (
        db.query(Usuario)
        .filter(Usuario.es_admin == True, Usuario.activo == True)  # noqa: E712
        .count()
    )
    if admins_activos == 0:
        db.rollback()
        result.errores.append((
            0,
            "El import dejaría el sistema sin ningún administrador activo. Abortado.",
        ))
        return result

    db.commit()
    result.commit_ok = True
    return result


# --- Horario de apertura de centros ------------------------------------

_HORARIO_CABECERAS = [
    "empresa_cif", "centro", "dia_semana", "cerrado",
    "apertura", "inicio_pausa", "fin_pausa", "cierre",
]


def exportar_horarios(db: Session) -> bytes:
    filas_db = (
        db.query(CentroHorarioApertura)
        .join(CentroTrabajo).join(Empresa)
        .order_by(Empresa.nombre, CentroTrabajo.nombre, CentroHorarioApertura.dia_semana)
        .all()
    )
    wb, ws = _crear_libro("Horarios", _HORARIO_CABECERAS)
    datos = []
    for h in filas_db:
        c = h.centro
        fila = [
            c.empresa.cif if c and c.empresa else "",
            _centro_referencia(c),
            _DIA_CODIGOS[h.dia_semana],
            _bool_to_excel(h.cerrado),
            h.apertura or "",
            h.inicio_pausa or "",
            h.fin_pausa or "",
            h.cierre or "",
        ]
        ws.append(fila)
        datos.append(fila)
    _autosize(ws, datos, _HORARIO_CABECERAS)
    return _serializar(wb)


# --- Soporte del formato "Google Business" --------------------------------
#
# Google Business exporta las ubicaciones en un Excel propio (una fila por
# tienda, ~80 columnas, horario de cada día como texto). Su "Código de tienda"
# no se corresponde con el codigo_centro de la app, así que cada tienda se
# empareja con un centro existente por su dirección (calle + municipio + CP).

# Cabecera (normalizada: minúsculas y sin acentos) → día de semana 0..6.
_GOOGLE_DIAS = {
    "horario de lunes": 0,
    "horario de martes": 1,
    "horario de miercoles": 2,
    "horario de jueves": 3,
    "horario de viernes": 4,
    "horario de sabado": 5,
    "horario de domingo": 6,
}

# Prefijos de tipo de vía (ya sin barras ni puntos) que se descartan al
# normalizar una dirección para el emparejamiento.
_PREFIJOS_VIA = {
    "calle", "c", "avenida", "avda", "avd", "av", "plaza", "pza", "pz", "pl",
    "paseo", "po", "ps", "carretera", "ctra", "cr", "poligono", "pol", "pg",
    "ronda", "rda", "camino", "cmno", "travesia", "trav", "urbanizacion",
    "urb", "barrio", "bo", "lugar", "via",
}


def _strip_acentos(s: str) -> str:
    """Elimina tildes y otros diacríticos de una cadena."""
    return "".join(
        c for c in unicodedata.normalize("NFKD", s)
        if not unicodedata.combining(c)
    )


def _norm_cabecera(c: Any) -> str:
    """Cabecera normalizada: minúsculas, sin acentos y sin espacios sobrantes."""
    return _strip_acentos(_norm_str(c)).lower()


def _cabeceras_excel(file_bytes: bytes) -> list[str]:
    """Devuelve la primera fila (cabeceras) del Excel, o [] si no se puede leer."""
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        return []
    for row in wb.active.iter_rows(values_only=True):
        return [c if c is not None else "" for c in row]
    return []


def _es_formato_google(cabeceras: Iterable[Any]) -> bool:
    """True si las cabeceras corresponden al export de Google Business."""
    norm = {_norm_cabecera(c) for c in cabeceras}
    tiene_dias = len(norm & set(_GOOGLE_DIAS)) >= 3
    tiene_direccion = "direccion (linea 1)" in norm
    return tiene_dias and tiene_direccion


def _norm_direccion(s: Any) -> tuple[str, str]:
    """Normaliza una dirección para emparejar: (nombre_via, numero).

    Pasa a minúsculas sin acentos, descarta el prefijo de tipo de vía y separa
    el número de portal del nombre de la calle.
    """
    txt = _strip_acentos(_norm_str(s)).lower()
    txt = re.sub(r"[^a-z0-9\s]", " ", txt)
    tokens = txt.split()
    if not tokens:
        return "", ""
    if tokens[0] in _PREFIJOS_VIA:
        tokens = tokens[1:]
    numeros = [t for t in tokens if t.isdigit()]
    calle = " ".join(t for t in tokens if not t.isdigit())
    return calle, (numeros[-1] if numeros else "")


def _parse_horario_google(celda: Any) -> dict:
    """Convierte una celda de horario de Google a los campos del horario.

    Devuelve un dict con cerrado / apertura / cierre / inicio_pausa / fin_pausa.
    Lanza ValueError si el texto no se puede representar (3+ tramos, formato
    desconocido, etc.).
    """
    cerrado_dict = {
        "cerrado": True, "apertura": None, "cierre": None,
        "inicio_pausa": None, "fin_pausa": None,
    }
    s = _norm_str(celda)
    if not s:
        return cerrado_dict
    low = _strip_acentos(s).lower()
    if low in {"cerrado", "closed"}:
        return cerrado_dict
    if "24" in low and ("hora" in low or "hour" in low):
        return {
            "cerrado": False, "apertura": "00:00", "cierre": "23:59",
            "inicio_pausa": None, "fin_pausa": None,
        }
    s = s.replace("–", "-").replace("—", "-")
    tramos: list[tuple[str, str]] = []
    for trozo in s.split(","):
        trozo = trozo.strip()
        if not trozo:
            continue
        partes = trozo.split("-")
        if len(partes) != 2:
            raise ValueError(f"tramo horario no reconocido: {trozo!r}")
        ini = _to_horario(partes[0])
        fin = _to_horario(partes[1])
        if not ini or not fin:
            raise ValueError(f"tramo horario no reconocido: {trozo!r}")
        tramos.append((ini, fin))
    if len(tramos) == 1:
        return {
            "cerrado": False, "apertura": tramos[0][0], "cierre": tramos[0][1],
            "inicio_pausa": None, "fin_pausa": None,
        }
    if len(tramos) == 2:
        return {
            "cerrado": False,
            "apertura": tramos[0][0], "inicio_pausa": tramos[0][1],
            "fin_pausa": tramos[1][0], "cierre": tramos[1][1],
        }
    raise ValueError(
        f"el horario tiene {len(tramos)} tramos y solo se admiten 1 o 2: {s!r}"
    )


def _emparejar_centro(
    centros: list[CentroTrabajo], dir1: str, municipio: str, cp: str,
) -> tuple[CentroTrabajo | None, str]:
    """Empareja una tienda de Google con un centro por su dirección.

    Devuelve (centro, confianza), con confianza "alta" / "media" cuando hay
    emparejamiento, "ambigua" si dos centros empatan, o "" si no hay ninguno.
    """
    calle_g, num_g = _norm_direccion(dir1)
    if not calle_g:
        return None, ""
    cp_g = _norm_str(cp)
    muni_g = _strip_acentos(_norm_str(municipio)).lower()

    candidatos: list[tuple[float, float, CentroTrabajo]] = []
    for c in centros:
        calle_c, num_c = _norm_direccion(c.direccion)
        if not calle_c:
            continue
        ratio = difflib.SequenceMatcher(None, calle_g, calle_c).ratio()
        cp_ok = bool(cp_g) and cp_g == _norm_str(c.codigo_postal)
        muni_c = _strip_acentos(_norm_str(c.ciudad)).lower()
        muni_ok = bool(muni_g) and bool(muni_c) and (
            muni_g in muni_c or muni_c in muni_g
        )
        num_ok = bool(num_g) and num_g == num_c
        score = 0.6 * ratio + 0.2 * cp_ok + 0.1 * muni_ok + 0.1 * num_ok
        candidatos.append((score, ratio, c))

    if not candidatos:
        return None, ""
    candidatos.sort(key=lambda t: t[0], reverse=True)
    mejor_score, mejor_ratio, mejor = candidatos[0]
    segundo_score = candidatos[1][0] if len(candidatos) > 1 else 0.0

    if mejor_score < 0.55 or mejor_ratio < 0.5:
        return None, ""
    if mejor_score - segundo_score < 0.08:
        return None, "ambigua"
    if mejor_ratio >= 0.85:
        return mejor, "alta"
    return mejor, "media"


def _upsert_horario(
    db: Session, result: ImportResult, centro_id: int, dow: int, h: dict,
) -> None:
    """Crea o actualiza el horario de apertura de (centro, día)."""
    existente = (
        db.query(CentroHorarioApertura)
        .filter(
            CentroHorarioApertura.centro_id == centro_id,
            CentroHorarioApertura.dia_semana == dow,
        )
        .one_or_none()
    )
    if existente is None:
        db.add(CentroHorarioApertura(
            centro_id=centro_id, dia_semana=dow, cerrado=h["cerrado"],
            apertura=h["apertura"], cierre=h["cierre"],
            inicio_pausa=h["inicio_pausa"], fin_pausa=h["fin_pausa"],
        ))
        result.creados += 1
    else:
        existente.cerrado = h["cerrado"]
        existente.apertura = h["apertura"]
        existente.cierre = h["cierre"]
        existente.inicio_pausa = h["inicio_pausa"]
        existente.fin_pausa = h["fin_pausa"]
        result.actualizados += 1


def importar_horarios(
    db: Session, file_bytes: bytes, empresa_id: int | None = None,
) -> ImportResult:
    """Importa horarios de apertura.

    Detecta automáticamente el formato del fichero: el Excel propio de la app
    o el export de Google Business. `empresa_id` solo se usa en el segundo.
    """
    if _es_formato_google(_cabeceras_excel(file_bytes)):
        return _importar_horarios_google(db, file_bytes, empresa_id)
    return _importar_horarios_estandar(db, file_bytes)


def _importar_horarios_estandar(db: Session, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    filas, errores = _leer_filas(file_bytes, ["empresa_cif", "centro", "dia_semana"])
    result.errores.extend(errores)
    if errores:
        return result

    vistos: set[tuple[int, int]] = set()
    for fila in filas:
        n = fila["__row__"]
        try:
            empresa = db.query(Empresa).filter(
                Empresa.cif == _norm_upper(fila.get("empresa_cif"))).one_or_none()
            if empresa is None:
                raise ValueError("No existe la empresa indicada.")
            centro = _resolver_centro(db, empresa.id, _norm_str(fila.get("centro")))
            dow = _to_dow(fila.get("dia_semana"))
            if (centro.id, dow) in vistos:
                raise ValueError("Día de ese centro duplicado en el fichero.")
            vistos.add((centro.id, dow))
            cerrado = bool(_to_bool(fila.get("cerrado"), default=False))
            apertura = None if cerrado else _to_horario(fila.get("apertura"))
            cierre = None if cerrado else _to_horario(fila.get("cierre"))
            ini_pausa = None if cerrado else _to_horario(fila.get("inicio_pausa"))
            fin_pausa = None if cerrado else _to_horario(fila.get("fin_pausa"))
            _upsert_horario(db, result, centro.id, dow, {
                "cerrado": cerrado, "apertura": apertura, "cierre": cierre,
                "inicio_pausa": ini_pausa, "fin_pausa": fin_pausa,
            })
        except ValueError as exc:
            result.errores.append((n, str(exc)))

    if result.errores:
        db.rollback()
        return result
    db.commit()
    result.commit_ok = True
    return result


def _importar_horarios_google(
    db: Session, file_bytes: bytes, empresa_id: int | None,
) -> ImportResult:
    result = ImportResult()
    if not empresa_id:
        result.errores.append((
            0,
            "Selecciona la empresa de destino para importar el formato "
            "Google Business.",
        ))
        return result
    empresa = db.get(Empresa, empresa_id)
    if empresa is None:
        result.errores.append((0, "La empresa seleccionada no existe."))
        return result

    filas, errores = _leer_filas(file_bytes, [])
    result.errores.extend(errores)
    if errores:
        return result

    centros = (
        db.query(CentroTrabajo)
        .filter(
            CentroTrabajo.empresa_id == empresa_id,
            CentroTrabajo.activo == True,  # noqa: E712
        )
        .all()
    )
    if not centros:
        result.errores.append((
            0, f"La empresa «{empresa.nombre}» no tiene centros activos.",
        ))
        return result

    centros_usados: set[int] = set()
    for fila_raw in filas:
        n = fila_raw["__row__"]
        # Reescribe las claves sin acentos para acceder a las columnas Google.
        fila = {
            (_norm_cabecera(k) if isinstance(k, str) else k): v
            for k, v in fila_raw.items()
        }
        dir1 = _norm_str(fila.get("direccion (linea 1)"))
        municipio = _norm_str(fila.get("municipio"))
        cp = _norm_str(fila.get("codigo postal"))
        if not dir1:
            result.avisos.append((n, "Tienda sin dirección; no se empareja."))
            continue

        centro, confianza = _emparejar_centro(centros, dir1, municipio, cp)
        if centro is None:
            ubic = f"«{dir1}»" + (f" ({municipio})" if municipio else "")
            if confianza == "ambigua":
                result.avisos.append((
                    n, f"{ubic} coincide con varios centros; no se empareja.",
                ))
            else:
                result.avisos.append((
                    n, f"No se encontró ningún centro para {ubic}.",
                ))
            continue
        if centro.id in centros_usados:
            result.avisos.append((
                n,
                f"El centro «{centro.nombre}» ya se emparejó con otra tienda "
                f"del fichero; se omite «{dir1}».",
            ))
            continue
        centros_usados.add(centro.id)
        if confianza == "media":
            result.avisos.append((
                n,
                f"«{dir1}» se ha emparejado con el centro «{centro.nombre}» "
                f"— revisa que sea correcto.",
            ))

        for cabecera, dow in _GOOGLE_DIAS.items():
            try:
                horario = _parse_horario_google(fila.get(cabecera))
            except ValueError as exc:
                result.avisos.append((
                    n, f"{centro.nombre} — {_DIA_CODIGOS[dow]}: {exc}",
                ))
                continue
            _upsert_horario(db, result, centro.id, dow, horario)

    if result.errores:
        db.rollback()
        return result
    db.commit()
    result.commit_ok = True
    return result


# --- Plantillas de turno -----------------------------------------------

_TURNO_CABECERAS = [
    "empresa_cif", "centro", "nombre", "hora_inicio", "hora_fin",
    "inicio_pausa", "fin_pausa", "activo", "orden",
]


def exportar_turnos(db: Session) -> bytes:
    filas_db = (
        db.query(PlantillaTurno)
        .join(CentroTrabajo).join(Empresa)
        .order_by(Empresa.nombre, CentroTrabajo.nombre, PlantillaTurno.orden)
        .all()
    )
    wb, ws = _crear_libro("Turnos", _TURNO_CABECERAS)
    datos = []
    for t in filas_db:
        c = t.centro
        fila = [
            c.empresa.cif if c and c.empresa else "",
            _centro_referencia(c),
            t.nombre,
            t.hora_inicio,
            t.hora_fin,
            t.inicio_pausa or "",
            t.fin_pausa or "",
            _bool_to_excel(t.activo),
            t.orden,
        ]
        ws.append(fila)
        datos.append(fila)
    _autosize(ws, datos, _TURNO_CABECERAS)
    return _serializar(wb)


def importar_turnos(db: Session, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    filas, errores = _leer_filas(
        file_bytes, ["empresa_cif", "centro", "nombre", "hora_inicio", "hora_fin"])
    result.errores.extend(errores)
    if errores:
        return result

    vistos: set[tuple[int, str]] = set()
    for fila in filas:
        n = fila["__row__"]
        try:
            empresa = db.query(Empresa).filter(
                Empresa.cif == _norm_upper(fila.get("empresa_cif"))).one_or_none()
            if empresa is None:
                raise ValueError("No existe la empresa indicada.")
            centro = _resolver_centro(db, empresa.id, _norm_str(fila.get("centro")))
            nombre = _norm_str(fila.get("nombre"))
            if not nombre:
                raise ValueError("El nombre del turno es obligatorio.")
            clave = (centro.id, nombre.lower())
            if clave in vistos:
                raise ValueError(f"Turno {nombre!r} duplicado en el fichero.")
            vistos.add(clave)
            hi = _to_horario(fila.get("hora_inicio"))
            hf = _to_horario(fila.get("hora_fin"))
            if not hi or not hf:
                raise ValueError("hora_inicio y hora_fin son obligatorias.")
            ip = _to_horario(fila.get("inicio_pausa"))
            fp = _to_horario(fila.get("fin_pausa"))
            orden_v = _to_numero(fila.get("orden"))
            existente = (
                db.query(PlantillaTurno)
                .filter(
                    PlantillaTurno.centro_id == centro.id,
                    PlantillaTurno.nombre == nombre,
                )
                .one_or_none()
            )
            activo = _to_bool(
                fila.get("activo"),
                default=True if existente is None else existente.activo,
            )
            if existente is None:
                db.add(PlantillaTurno(
                    centro_id=centro.id, nombre=nombre,
                    hora_inicio=hi, hora_fin=hf, inicio_pausa=ip, fin_pausa=fp,
                    activo=bool(activo), orden=int(orden_v or 0),
                ))
                result.creados += 1
            else:
                existente.hora_inicio = hi
                existente.hora_fin = hf
                existente.inicio_pausa = ip
                existente.fin_pausa = fp
                existente.activo = bool(activo)
                existente.orden = int(orden_v or 0)
                result.actualizados += 1
        except ValueError as exc:
            result.errores.append((n, str(exc)))

    if result.errores:
        db.rollback()
        return result
    db.commit()
    result.commit_ok = True
    return result


# --- Dotación -----------------------------------------------------------

_DOTACION_CABECERAS = [
    "empresa_cif", "centro", "turno", "dia_semana", "minimo", "maximo",
]


def exportar_dotacion(db: Session) -> bytes:
    filas_db = (
        db.query(DotacionCentro)
        .join(CentroTrabajo, CentroTrabajo.id == DotacionCentro.centro_id)
        .join(Empresa)
        .order_by(Empresa.nombre, CentroTrabajo.nombre, DotacionCentro.dia_semana)
        .all()
    )
    wb, ws = _crear_libro("Dotacion", _DOTACION_CABECERAS)
    datos = []
    for d in filas_db:
        c = d.centro
        fila = [
            c.empresa.cif if c and c.empresa else "",
            _centro_referencia(c),
            d.plantilla_turno.nombre if d.plantilla_turno else "",
            _DIA_CODIGOS[d.dia_semana],
            d.minimo,
            d.maximo,
        ]
        ws.append(fila)
        datos.append(fila)
    _autosize(ws, datos, _DOTACION_CABECERAS)
    return _serializar(wb)


def importar_dotacion(db: Session, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    filas, errores = _leer_filas(
        file_bytes, ["empresa_cif", "centro", "turno", "dia_semana"])
    result.errores.extend(errores)
    if errores:
        return result

    vistos: set[tuple[int, int, int]] = set()
    for fila in filas:
        n = fila["__row__"]
        try:
            empresa = db.query(Empresa).filter(
                Empresa.cif == _norm_upper(fila.get("empresa_cif"))).one_or_none()
            if empresa is None:
                raise ValueError("No existe la empresa indicada.")
            centro = _resolver_centro(db, empresa.id, _norm_str(fila.get("centro")))
            nombre_turno = _norm_str(fila.get("turno"))
            turno = (
                db.query(PlantillaTurno)
                .filter(
                    PlantillaTurno.centro_id == centro.id,
                    PlantillaTurno.nombre == nombre_turno,
                )
                .one_or_none()
            )
            if turno is None:
                raise ValueError(f"No existe el turno {nombre_turno!r} en el centro.")
            dow = _to_dow(fila.get("dia_semana"))
            clave = (centro.id, turno.id, dow)
            if clave in vistos:
                raise ValueError("Combinación turno/día duplicada en el fichero.")
            vistos.add(clave)
            minimo = int(_to_numero(fila.get("minimo")) or 0)
            maximo = int(_to_numero(fila.get("maximo")) or 0)
            if minimo < 0 or maximo < 0:
                raise ValueError("La dotación no puede ser negativa.")
            if maximo < minimo:
                raise ValueError("El máximo no puede ser menor que el mínimo.")
            existente = (
                db.query(DotacionCentro)
                .filter(
                    DotacionCentro.centro_id == centro.id,
                    DotacionCentro.plantilla_turno_id == turno.id,
                    DotacionCentro.dia_semana == dow,
                )
                .one_or_none()
            )
            if existente is None:
                db.add(DotacionCentro(
                    centro_id=centro.id, plantilla_turno_id=turno.id,
                    dia_semana=dow, minimo=minimo, maximo=maximo,
                ))
                result.creados += 1
            else:
                existente.minimo = minimo
                existente.maximo = maximo
                result.actualizados += 1
        except ValueError as exc:
            result.errores.append((n, str(exc)))

    if result.errores:
        db.rollback()
        return result
    db.commit()
    result.commit_ok = True
    return result
